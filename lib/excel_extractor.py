import openpyxl
import pandas as pd
import json
import io
from typing import Dict, Any, Optional, List, Tuple
try:
    from .source_tracker import SourceTracker
except ImportError:
    from source_tracker import SourceTracker

class ExcelExtractor:
    def __init__(self, source_tracker: Optional[SourceTracker] = None):
        """Initialize with optional source tracker for enhanced attribution"""
        self.source_tracker = source_tracker
        
    def extract_with_coordinates(self, file_path):
        """
        Extract data maintaining cell references for traceability.
        Returns: {
            'sheet_name': {
                'data': DataFrame as dict,
                'key_metrics': {
                    'metric_name': {
                        'value': actual_value,
                        'cell': 'B15',
                        'formula': '=SUM(B1:B14)'
                    }
                },
                'tables': [
                    {
                        'range': 'A1:D10',
                        'title': 'Financial Summary',
                        'data': [[...]]
                    }
                ]
            }
        }
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=False)
            result = {}
            
            # Register document with source tracker if available
            document_id = None
            if self.source_tracker:
                document_id = self.source_tracker.register_document(
                    file_path, 
                    'excel',
                    {'workbook_sheets': workbook.sheetnames, 'sheet_count': len(workbook.sheetnames)}
                )
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = self._extract_sheet_data(sheet, sheet_name, file_path, document_id)
                result[sheet_name] = sheet_data
            
            # Add enhanced attribution data if source tracker is available
            if self.source_tracker and document_id:
                result['_attribution'] = {
                    'document_id': document_id,
                    'source_tracker_data': self.source_tracker.export_attribution_data()
                }
            
            return result
        except Exception as e:
            return {'error': f'Failed to extract Excel data: {str(e)}'}
    
    def extract_from_bytes(self, excel_bytes, filename="document.xlsx"):
        """Similar to PDF extractor's pattern with enhanced source tracking"""
        try:
            # Create a BytesIO object from the bytes
            excel_file = io.BytesIO(excel_bytes)
            workbook = openpyxl.load_workbook(excel_file, data_only=False)
            result = {'filename': filename, 'sheets': {}}
            
            # Register document with source tracker if available
            document_id = None
            if self.source_tracker:
                document_id = self.source_tracker.register_document(
                    filename, 
                    'excel',
                    {'workbook_sheets': workbook.sheetnames, 'sheet_count': len(workbook.sheetnames)}
                )
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = self._extract_sheet_data(sheet, sheet_name, filename, document_id)
                result['sheets'][sheet_name] = sheet_data
            
            # Add enhanced attribution data if source tracker is available
            if self.source_tracker and document_id:
                result['_attribution'] = {
                    'document_id': document_id,
                    'source_tracker_data': self.source_tracker.export_attribution_data()
                }
            
            return result
        except Exception as e:
            return {'error': f'Failed to extract Excel data from bytes: {str(e)}'}
    
    def _extract_sheet_data(self, sheet, sheet_name, filename=None, document_id=None):
        """Extract data from a single sheet with enhanced source tracking"""
        sheet_data = {
            'data': {},
            'key_metrics': {},
            'tables': [],
            'formulas': {},
            'data_point_ids': {}  # Maps cell references to data point IDs
        }
        
        # Get all non-empty cells
        data_rows = []
        max_row = min(sheet.max_row, 100)  # Limit to first 100 rows for performance
        max_col = min(sheet.max_column, 50)  # Limit to first 50 columns
        
        for row in range(1, max_row + 1):
            row_data = []
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                cell_ref = cell.coordinate
                
                # Store cell value and formula if exists
                value = cell.value
                formula = cell.formula if hasattr(cell, 'formula') else None
                
                # Track data point with source tracker if available and value is significant
                data_point_id = None
                if self.source_tracker and document_id and value is not None:
                    # Get surrounding context for better attribution
                    context = self._get_cell_context(sheet, row, col)
                    
                    # Determine confidence based on data type and context
                    confidence = self._calculate_extraction_confidence(value, formula, context)
                    
                    # Track the data point
                    data_point_id = self.source_tracker.track_data_point(
                        value=value,
                        document_id=document_id,
                        location_details={
                            'page_or_sheet': sheet_name,
                            'cell_or_section': cell_ref,
                            'table_name': context.get('table_name'),
                            'coordinates': {'row': row, 'col': col},
                            'extraction_method': 'openpyxl'
                        },
                        confidence=confidence,
                        context=context.get('description'),
                        formula=formula
                    )
                    
                    sheet_data['data_point_ids'][cell_ref] = data_point_id
                
                if formula and formula.startswith('='):
                    sheet_data['formulas'][cell_ref] = {
                        'formula': formula,
                        'value': value,
                        'data_point_id': data_point_id
                    }
                
                row_data.append({
                    'value': value,
                    'cell': cell_ref,
                    'formula': formula,
                    'data_point_id': data_point_id
                })
            
            data_rows.append(row_data)
        
        sheet_data['data'] = data_rows
        
        # Try to identify key metrics (numbers with formulas or obvious patterns)
        self._identify_key_metrics(sheet, sheet_data, max_row, max_col, document_id)
        
        # Try to identify table structures
        self._identify_tables(sheet_data, data_rows)
        
        return sheet_data
    
    def _identify_key_metrics(self, sheet, sheet_data, max_row, max_col, document_id=None):
        """Identify potential key financial metrics with enhanced source tracking"""
        metrics = {}
        
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                cell_ref = cell.coordinate
                
                if cell.value is not None:
                    # Look for cells with formulas that contain numbers
                    if hasattr(cell, 'formula') and cell.formula and cell.formula.startswith('='):
                        if isinstance(cell.value, (int, float)) and cell.value != 0:
                            # Try to guess metric name from nearby cells
                            metric_name = self._guess_metric_name(sheet, row, col)
                            if metric_name:
                                # Get data point ID if available
                                data_point_id = sheet_data.get('data_point_ids', {}).get(cell_ref)
                                
                                metrics[metric_name] = {
                                    'value': cell.value,
                                    'cell': cell.coordinate,
                                    'formula': cell.formula,
                                    'data_point_id': data_point_id
                                }
                    
                    # Look for large numbers that might be financial figures
                    elif isinstance(cell.value, (int, float)) and abs(cell.value) > 1000:
                        metric_name = self._guess_metric_name(sheet, row, col)
                        if metric_name:
                            # Get data point ID if available
                            data_point_id = sheet_data.get('data_point_ids', {}).get(cell_ref)
                            
                            metrics[metric_name] = {
                                'value': cell.value,
                                'cell': cell.coordinate,
                                'formula': None,
                                'data_point_id': data_point_id
                            }
        
        sheet_data['key_metrics'] = metrics
    
    def _guess_metric_name(self, sheet, row, col):
        """Try to guess the metric name from nearby cells"""
        # Look left and up for text that might be a label
        for offset in range(1, 4):
            # Check left
            if col - offset > 0:
                left_cell = sheet.cell(row=row, column=col - offset)
                if isinstance(left_cell.value, str) and len(left_cell.value) > 2:
                    return left_cell.value.strip()
            
            # Check up
            if row - offset > 0:
                up_cell = sheet.cell(row=row - offset, column=col)
                if isinstance(up_cell.value, str) and len(up_cell.value) > 2:
                    return up_cell.value.strip()
        
        return None
    
    def _identify_tables(self, sheet_data, data_rows):
        """Identify rectangular table structures in the data"""
        tables = []
        
        # Simple heuristic: look for rectangular regions with headers
        if len(data_rows) > 2:
            # Check if first row looks like headers (mostly text)
            first_row = data_rows[0]
            header_cells = [cell for cell in first_row if isinstance(cell['value'], str) and cell['value']]
            
            if len(header_cells) >= 2:  # At least 2 headers
                # Count consecutive rows with data
                data_row_count = 0
                for i in range(1, len(data_rows)):
                    row = data_rows[i]
                    non_empty = [cell for cell in row if cell['value'] is not None]
                    if len(non_empty) >= 2:
                        data_row_count += 1
                    else:
                        break
                
                if data_row_count >= 2:  # At least 2 data rows
                    table_range = f"A1:{chr(65 + len(header_cells) - 1)}{data_row_count + 1}"
                    tables.append({
                        'range': table_range,
                        'title': 'Data Table',
                        'headers': [cell['value'] for cell in header_cells],
                        'data_rows': data_row_count
                    })
        
        sheet_data['tables'] = tables
    
    def _get_cell_context(self, sheet, row: int, col: int) -> Dict[str, Any]:
        """Get contextual information around a cell for better source attribution"""
        context = {
            'description': '',
            'table_name': None,
            'nearby_labels': [],
            'is_header': False,
            'data_type_hint': None
        }
        
        # Check if this cell might be part of a table header
        cell_value = sheet.cell(row=row, column=col).value
        if isinstance(cell_value, str) and len(cell_value) > 2:
            # Check if next row has numeric data (indicating this might be a header)
            if row + 1 <= sheet.max_row:
                next_cell = sheet.cell(row=row + 1, column=col).value
                if isinstance(next_cell, (int, float)):
                    context['is_header'] = True
                    context['table_name'] = cell_value
        
        # Collect nearby text labels for context
        for r_offset in [-2, -1, 1, 2]:
            for c_offset in [-2, -1, 1, 2]:
                if r_offset == 0 and c_offset == 0:
                    continue
                
                try:
                    nearby_row = row + r_offset
                    nearby_col = col + c_offset
                    if nearby_row > 0 and nearby_col > 0:
                        nearby_cell = sheet.cell(row=nearby_row, column=nearby_col)
                        if isinstance(nearby_cell.value, str) and len(nearby_cell.value.strip()) > 2:
                            context['nearby_labels'].append({
                                'text': nearby_cell.value.strip(),
                                'position': f'{r_offset},{c_offset}',
                                'cell': nearby_cell.coordinate
                            })
                except:
                    continue
        
        # Create description from labels
        if context['nearby_labels']:
            primary_labels = [label['text'] for label in context['nearby_labels'][:3]]
            context['description'] = f"Near: {', '.join(primary_labels)}"
        
        # Determine data type hint
        if isinstance(cell_value, (int, float)):
            if abs(cell_value) > 1000000:
                context['data_type_hint'] = 'large_financial'
            elif abs(cell_value) > 1000:
                context['data_type_hint'] = 'medium_financial'
            elif 0 <= cell_value <= 1:
                context['data_type_hint'] = 'percentage_decimal'
            else:
                context['data_type_hint'] = 'numeric'
        elif isinstance(cell_value, str):
            if '%' in cell_value:
                context['data_type_hint'] = 'percentage_text'
            elif '$' in cell_value or '€' in cell_value or '£' in cell_value:
                context['data_type_hint'] = 'currency_text'
            else:
                context['data_type_hint'] = 'text'
        
        return context
    
    def _calculate_extraction_confidence(self, value: Any, formula: Optional[str], 
                                       context: Dict[str, Any]) -> float:
        """Calculate confidence score for data extraction based on multiple factors"""
        confidence = 0.8  # Base confidence
        
        # Formula-based adjustments
        if formula and formula.startswith('='):
            confidence += 0.1  # Formulas are generally more reliable
            
            # Complex formulas get higher confidence
            if any(func in formula.upper() for func in ['SUM', 'AVERAGE', 'COUNT', 'IF']):
                confidence += 0.05
        
        # Data type adjustments
        if isinstance(value, (int, float)):
            confidence += 0.05  # Numeric values are generally more reliable
            
            # Large numbers in financial context get higher confidence
            if context.get('data_type_hint') in ['large_financial', 'medium_financial']:
                confidence += 0.05
        
        # Context-based adjustments
        if context.get('is_header'):
            confidence -= 0.1  # Headers are less likely to be data points
        
        if context.get('nearby_labels'):
            confidence += 0.05  # Having context labels increases confidence
        
        # Table context
        if context.get('table_name'):
            confidence += 0.1  # Being part of a named table increases confidence
        
        # Cap confidence at 1.0
        return min(confidence, 1.0)
