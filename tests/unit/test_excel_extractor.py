"""
Comprehensive tests for Excel extractor with formula parsing and cell referencing
Tests cover various spreadsheet formats, formulas, multiple sheets, and source attribution
"""

import unittest
from unittest.mock import Mock, patch, MagicMock
import sys
import os
import io
import openpyxl
from openpyxl import Workbook

# Add the lib directory to the path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..', 'lib'))

from excel_extractor import ExcelExtractor
from source_tracker import SourceTracker


class TestExcelExtractor(unittest.TestCase):
    """Test suite for Excel extractor functionality"""

    def setUp(self):
        """Set up test fixtures"""
        self.source_tracker = SourceTracker()
        self.excel_extractor = ExcelExtractor(source_tracker=self.source_tracker)
        
    def create_test_workbook(self):
        """Create a test Excel workbook for testing"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Financial Data"
        
        # Add headers
        ws['A1'] = "Metric"
        ws['B1'] = "Value"
        ws['C1'] = "Formula"
        
        # Add data with formulas
        ws['A2'] = "Revenue"
        ws['B2'] = 1000000
        ws['C2'] = "=B2"
        
        ws['A3'] = "Costs"
        ws['B3'] = 600000
        ws['C3'] = "=B3"
        
        ws['A4'] = "Profit"
        ws['B4'] = "=B2-B3"
        ws['C4'] = "=B2-B3"
        
        ws['A5'] = "Margin"
        ws['B5'] = "=B4/B2"
        ws['C5'] = "=(B2-B3)/B2"
        
        # Add a second sheet with table data
        ws2 = wb.create_sheet("Quarterly Data")
        quarters = ["Q1", "Q2", "Q3", "Q4"]
        revenue_data = [250000, 275000, 230000, 245000]
        
        # Headers
        ws2['A1'] = "Quarter"
        ws2['B1'] = "Revenue"
        ws2['C1'] = "Growth"
        
        # Data
        for i, (quarter, revenue) in enumerate(zip(quarters, revenue_data), start=2):
            ws2[f'A{i}'] = quarter
            ws2[f'B{i}'] = revenue
            if i > 2:  # Add growth calculation starting from Q2
                ws2[f'C{i}'] = f"=(B{i}-B{i-1})/B{i-1}"
        
        # Add total
        ws2['A6'] = "Total"
        ws2['B6'] = "=SUM(B2:B5)"
        
        return wb

    def test_init_without_source_tracker(self):
        """Test initialization without source tracker"""
        extractor = ExcelExtractor()
        self.assertIsNone(extractor.source_tracker)

    def test_init_with_source_tracker(self):
        """Test initialization with source tracker"""
        self.assertIsNotNone(self.excel_extractor.source_tracker)
        self.assertEqual(self.excel_extractor.source_tracker, self.source_tracker)

    def test_extract_from_bytes_success(self):
        """Test successful Excel extraction from bytes"""
        # Create test workbook
        wb = self.create_test_workbook()
        
        # Save to bytes
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        result = self.excel_extractor.extract_from_bytes(excel_bytes.getvalue(), "test.xlsx")
        
        # Verify result structure
        self.assertIn('filename', result)
        self.assertIn('sheets', result)
        self.assertEqual(result['filename'], "test.xlsx")
        
        # Verify sheets were extracted
        self.assertIn('Financial Data', result['sheets'])
        self.assertIn('Quarterly Data', result['sheets'])
        
        # Verify sheet data structure
        financial_sheet = result['sheets']['Financial Data']
        self.assertIn('data', financial_sheet)
        self.assertIn('key_metrics', financial_sheet)
        self.assertIn('tables', financial_sheet)
        self.assertIn('formulas', financial_sheet)

    def test_extract_from_bytes_with_source_tracking(self):
        """Test Excel extraction with source tracking enabled"""
        wb = self.create_test_workbook()
        
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        result = self.excel_extractor.extract_from_bytes(excel_bytes.getvalue(), "test.xlsx")
        
        # Verify attribution data exists
        self.assertIn('_attribution', result)
        self.assertIn('document_id', result['_attribution'])
        self.assertIn('source_tracker_data', result['_attribution'])
        
        # Verify data points were tracked
        self.assertTrue(len(self.source_tracker.data_points) > 0)
        
        # Verify sheet data has data point IDs
        financial_sheet = result['sheets']['Financial Data']
        self.assertIn('data_point_ids', financial_sheet)

    def test_extract_from_bytes_error_handling(self):
        """Test error handling for invalid Excel data"""
        invalid_bytes = b"not_an_excel_file"
        result = self.excel_extractor.extract_from_bytes(invalid_bytes, "invalid.xlsx")
        
        self.assertIn('error', result)
        self.assertIn('Failed to extract Excel data from bytes', result['error'])

    def test_extract_sheet_data_comprehensive(self):
        """Test comprehensive sheet data extraction"""
        wb = self.create_test_workbook()
        sheet = wb['Financial Data']
        
        # Register document for source tracking
        doc_id = self.source_tracker.register_document("test.xlsx", "excel")
        
        sheet_data = self.excel_extractor._extract_sheet_data(sheet, "Financial Data", "test.xlsx", doc_id)
        
        # Verify data structure
        self.assertIn('data', sheet_data)
        self.assertIn('key_metrics', sheet_data)
        self.assertIn('tables', sheet_data)
        self.assertIn('formulas', sheet_data)
        self.assertIn('data_point_ids', sheet_data)
        
        # Verify formulas were captured
        self.assertTrue(len(sheet_data['formulas']) > 0)
        self.assertIn('B4', sheet_data['formulas'])  # Profit formula
        self.assertEqual(sheet_data['formulas']['B4']['formula'], '=B2-B3')
        
        # Verify data point tracking
        self.assertTrue(len(sheet_data['data_point_ids']) > 0)
        self.assertTrue(len(self.source_tracker.data_points) > 0)

    def test_identify_key_metrics(self):
        """Test identification of key financial metrics"""
        wb = self.create_test_workbook()
        sheet = wb['Financial Data']
        
        sheet_data = {
            'data': [],
            'key_metrics': {},
            'tables': [],
            'formulas': {},
            'data_point_ids': {}
        }
        
        self.excel_extractor._identify_key_metrics(sheet, sheet_data, 10, 10)
        
        # Verify key metrics were identified
        self.assertTrue(len(sheet_data['key_metrics']) > 0)
        
        # Check for specific metrics
        metric_names = list(sheet_data['key_metrics'].keys())
        
        # Should find Revenue, Profit, etc.
        self.assertTrue(any('revenue' in name.lower() for name in metric_names))

    def test_guess_metric_name(self):
        """Test metric name guessing from nearby cells"""
        wb = Workbook()
        ws = wb.active
        
        # Set up test data
        ws['A1'] = "Total Revenue"
        ws['B1'] = 1000000
        
        ws['A2'] = "Net Profit"
        ws['B2'] = 250000
        
        # Test left cell lookup
        metric_name = self.excel_extractor._guess_metric_name(ws, 1, 2)  # B1
        self.assertEqual(metric_name, "Total Revenue")
        
        metric_name = self.excel_extractor._guess_metric_name(ws, 2, 2)  # B2
        self.assertEqual(metric_name, "Net Profit")
        
        # Test no label case
        ws['C3'] = 500
        metric_name = self.excel_extractor._guess_metric_name(ws, 3, 3)  # C3
        self.assertIsNone(metric_name)

    def test_identify_tables(self):
        """Test identification of table structures"""
        # Create test data that looks like a table
        data_rows = [
            # Headers row
            [
                {'value': 'Product', 'cell': 'A1', 'formula': None, 'data_point_id': None},
                {'value': 'Revenue', 'cell': 'B1', 'formula': None, 'data_point_id': None},
                {'value': 'Profit', 'cell': 'C1', 'formula': None, 'data_point_id': None}
            ],
            # Data rows
            [
                {'value': 'Product A', 'cell': 'A2', 'formula': None, 'data_point_id': None},
                {'value': 100000, 'cell': 'B2', 'formula': None, 'data_point_id': None},
                {'value': 25000, 'cell': 'C2', 'formula': None, 'data_point_id': None}
            ],
            [
                {'value': 'Product B', 'cell': 'A3', 'formula': None, 'data_point_id': None},
                {'value': 150000, 'cell': 'B3', 'formula': None, 'data_point_id': None},
                {'value': 35000, 'cell': 'C3', 'formula': None, 'data_point_id': None}
            ]
        ]
        
        sheet_data = {'tables': []}
        self.excel_extractor._identify_tables(sheet_data, data_rows)
        
        # Verify table was identified
        self.assertEqual(len(sheet_data['tables']), 1)
        table = sheet_data['tables'][0]
        
        self.assertEqual(table['range'], 'A1:C3')
        self.assertEqual(table['headers'], ['Product', 'Revenue', 'Profit'])
        self.assertEqual(table['data_rows'], 2)

    def test_get_cell_context(self):
        """Test cell context extraction for source attribution"""
        wb = Workbook()
        ws = wb.active
        
        # Set up context scenario
        ws['A1'] = "Financial Summary"
        ws['A2'] = "Revenue"
        ws['B2'] = 1000000  # Target cell
        ws['A3'] = "Costs"
        ws['B3'] = 750000
        
        context = self.excel_extractor._get_cell_context(ws, 2, 2)  # B2
        
        # Verify context structure
        self.assertIn('description', context)
        self.assertIn('table_name', context)
        self.assertIn('nearby_labels', context)
        self.assertIn('is_header', context)
        self.assertIn('data_type_hint', context)
        
        # Verify context content
        self.assertFalse(context['is_header'])  # Numeric cell shouldn't be header
        self.assertEqual(context['data_type_hint'], 'financial_large')
        self.assertTrue(len(context['nearby_labels']) > 0)

    def test_calculate_extraction_confidence(self):
        """Test confidence calculation for data extraction"""
        # Test high confidence case (formula with good context)
        context = {
            'is_header': False,
            'nearby_labels': [{'text': 'Total Revenue', 'position': '-1,0'}],
            'table_name': 'Financial Summary',
            'data_type_hint': 'large_financial'
        }
        
        confidence = self.excel_extractor._calculate_extraction_confidence(
            1000000, "=SUM(B1:B10)", context
        )
        
        self.assertGreater(confidence, 0.9)
        
        # Test lower confidence case (header cell)
        context['is_header'] = True
        confidence = self.excel_extractor._calculate_extraction_confidence(
            "Revenue", None, context
        )
        
        self.assertLess(confidence, 0.9)

    def test_multiple_sheets_extraction(self):
        """Test extraction from multiple sheets"""
        wb = self.create_test_workbook()
        
        # Add a third sheet
        ws3 = wb.create_sheet("Summary")
        ws3['A1'] = "Summary Data"
        ws3['A2'] = "Total Revenue"
        ws3['B2'] = "='Financial Data'!B2+'Quarterly Data'!B6"
        
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        result = self.excel_extractor.extract_from_bytes(excel_bytes.getvalue(), "multi_sheet.xlsx")
        
        # Verify all sheets were extracted
        self.assertEqual(len(result['sheets']), 3)
        self.assertIn('Financial Data', result['sheets'])
        self.assertIn('Quarterly Data', result['sheets'])
        self.assertIn('Summary', result['sheets'])
        
        # Verify cross-sheet formula
        summary_sheet = result['sheets']['Summary']
        self.assertTrue(any('Financial Data' in str(formula.get('formula', '')) 
                          for formula in summary_sheet['formulas'].values()))

    def test_formula_types_detection(self):
        """Test detection of different formula types"""
        wb = Workbook()
        ws = wb.active
        
        # Different formula types
        ws['A1'] = "=SUM(B1:B10)"
        ws['A2'] = "=AVERAGE(C1:C10)"
        ws['A3'] = "=COUNT(D1:D10)"
        ws['A4'] = "=IF(E1>100,E1*0.1,0)"
        ws['A5'] = "=B1*C1"
        ws['A6'] = "=VLOOKUP(F1,A:B,2,FALSE)"
        
        # Register document
        doc_id = self.source_tracker.register_document("formulas.xlsx", "excel")
        
        sheet_data = self.excel_extractor._extract_sheet_data(ws, "Sheet1", "formulas.xlsx", doc_id)
        
        # Verify formulas were captured
        formulas = sheet_data['formulas']
        self.assertTrue(len(formulas) >= 6)
        
        # Check specific formula functions
        formula_strings = [f['formula'] for f in formulas.values()]
        self.assertTrue(any('SUM' in f for f in formula_strings))
        self.assertTrue(any('AVERAGE' in f for f in formula_strings))
        self.assertTrue(any('IF' in f for f in formula_strings))

    def test_large_spreadsheet_performance(self):
        """Test performance with larger spreadsheets (within limits)"""
        wb = Workbook()
        ws = wb.active
        
        # Create a larger dataset (still within 100x50 limit)
        for row in range(1, 51):  # 50 rows
            for col in range(1, 21):  # 20 columns
                if row == 1:
                    ws.cell(row=row, column=col, value=f"Header_{col}")
                else:
                    if col <= 10:
                        ws.cell(row=row, column=col, value=row * col * 100)
                    else:
                        ws.cell(row=row, column=col, value=f"=A{row}*B{row}")
        
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        # This should complete without timeout
        result = self.excel_extractor.extract_from_bytes(excel_bytes.getvalue(), "large.xlsx")
        
        self.assertIn('sheets', result)
        self.assertNotIn('error', result)
        
        sheet_data = result['sheets']['Sheet']
        self.assertEqual(len(sheet_data['data']), 50)  # 50 rows
        self.assertTrue(len(sheet_data['formulas']) > 0)

    def test_edge_case_empty_cells(self):
        """Test handling of empty cells and sparse data"""
        wb = Workbook()
        ws = wb.active
        
        # Sparse data pattern
        ws['A1'] = "Header"
        ws['A3'] = "Data"
        ws['B5'] = 12345
        ws['D7'] = "=B5*2"  # Formula referencing sparse data
        
        sheet_data = self.excel_extractor._extract_sheet_data(ws, "Sparse", "sparse.xlsx")
        
        # Should handle sparse data gracefully
        self.assertIn('data', sheet_data)
        self.assertIn('formulas', sheet_data)
        
        # Verify formula was captured
        self.assertIn('D7', sheet_data['formulas'])
        self.assertEqual(sheet_data['formulas']['D7']['formula'], '=B5*2')

    def test_edge_case_formula_errors(self):
        """Test handling of formula errors and edge cases"""
        wb = Workbook()
        ws = wb.active
        
        # Set up cells that will create formula errors
        ws['A1'] = 10
        ws['A2'] = 0
        ws['A3'] = "=A1/A2"  # Division by zero
        ws['A4'] = "=VLOOKUP(X1,A:B,2,FALSE)"  # Lookup error
        ws['A5'] = "=SUM(Z:Z)"  # Reference to empty column
        
        sheet_data = self.excel_extractor._extract_sheet_data(ws, "Errors", "errors.xlsx")
        
        # Should handle errors gracefully
        self.assertIn('formulas', sheet_data)
        self.assertTrue(len(sheet_data['formulas']) > 0)

    def test_data_type_classification(self):
        """Test classification of different data types in cells"""
        wb = Workbook()
        ws = wb.active
        
        # Different data types
        ws['A1'] = 1000000  # Large number
        ws['A2'] = 0.25     # Decimal (could be percentage)
        ws['A3'] = "25%"    # Percentage string
        ws['A4'] = "$1,000" # Currency string
        ws['A5'] = "2023"   # Year
        ws['A6'] = "Regular text"
        
        # Register document for tracking
        doc_id = self.source_tracker.register_document("types.xlsx", "excel")
        
        sheet_data = self.excel_extractor._extract_sheet_data(ws, "Types", "types.xlsx", doc_id)
        
        # Verify data points were created with appropriate context
        self.assertTrue(len(self.source_tracker.data_points) > 0)
        
        # Check that different data types were recognized in context
        for data_point in self.source_tracker.data_points.values():
            self.assertIn(data_point.data_type, [
                'financial_large', 'numeric', 'text', 'percentage_decimal', 'year'
            ])


class TestExcelExtractorIntegration(unittest.TestCase):
    """Integration tests for Excel extractor with real-world scenarios"""

    def setUp(self):
        """Set up integration test fixtures"""
        self.source_tracker = SourceTracker()
        self.excel_extractor = ExcelExtractor(source_tracker=self.source_tracker)

    def test_financial_model_extraction(self):
        """Test extraction from a comprehensive financial model"""
        wb = Workbook()
        
        # Create Income Statement sheet
        income_sheet = wb.active
        income_sheet.title = "Income Statement"
        
        # Headers
        income_sheet['A1'] = "Income Statement"
        income_sheet['B1'] = "2023"
        income_sheet['C1'] = "2022"
        income_sheet['D1'] = "Change"
        
        # Revenue section
        income_sheet['A3'] = "Revenue"
        income_sheet['B3'] = 10000000
        income_sheet['C3'] = 8500000
        income_sheet['D3'] = "=B3-C3"
        
        # Cost section
        income_sheet['A4'] = "Cost of Sales"
        income_sheet['B4'] = 6000000
        income_sheet['C4'] = 5100000
        income_sheet['D4'] = "=B4-C4"
        
        # Gross Profit
        income_sheet['A5'] = "Gross Profit"
        income_sheet['B5'] = "=B3-B4"
        income_sheet['C5'] = "=C3-C4"
        income_sheet['D5'] = "=B5-C5"
        
        # Margins
        income_sheet['A7'] = "Gross Margin %"
        income_sheet['B7'] = "=B5/B3"
        income_sheet['C7'] = "=C5/C3"
        
        # Create Balance Sheet
        balance_sheet = wb.create_sheet("Balance Sheet")
        balance_sheet['A1'] = "Assets"
        balance_sheet['A2'] = "Cash"
        balance_sheet['B2'] = 2500000
        balance_sheet['A3'] = "Accounts Receivable"
        balance_sheet['B3'] = 1800000
        balance_sheet['A4'] = "Total Current Assets"
        balance_sheet['B4'] = "=B2+B3"
        
        # Create KPIs sheet with cross-references
        kpi_sheet = wb.create_sheet("KPIs")
        kpi_sheet['A1'] = "Key Performance Indicators"
        kpi_sheet['A2'] = "Revenue Growth"
        kpi_sheet['B2'] = "='Income Statement'!D3/'Income Statement'!C3"
        kpi_sheet['A3'] = "Cash Ratio"
        kpi_sheet['B3'] = "='Balance Sheet'!B2/'Income Statement'!B3"
        
        # Convert to bytes and extract
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        result = self.excel_extractor.extract_from_bytes(excel_bytes.getvalue(), "financial_model.xlsx")
        
        # Verify comprehensive extraction
        self.assertEqual(len(result['sheets']), 3)
        self.assertIn('Income Statement', result['sheets'])
        self.assertIn('Balance Sheet', result['sheets'])
        self.assertIn('KPIs', result['sheets'])
        
        # Verify complex formulas were captured
        income_formulas = result['sheets']['Income Statement']['formulas']
        self.assertTrue(any('B3-B4' in f['formula'] for f in income_formulas.values()))
        self.assertTrue(any('B5/B3' in f['formula'] for f in income_formulas.values()))
        
        # Verify cross-sheet references
        kpi_formulas = result['sheets']['KPIs']['formulas']
        cross_sheet_formulas = [f['formula'] for f in kpi_formulas.values() 
                               if 'Income Statement' in f['formula']]
        self.assertTrue(len(cross_sheet_formulas) > 0)

    def test_dashboard_with_charts_metadata(self):
        """Test extraction from dashboard-style spreadsheet"""
        wb = Workbook()
        
        # Summary dashboard
        summary = wb.active
        summary.title = "Dashboard"
        
        # Key metrics section
        summary['A1'] = "EXECUTIVE DASHBOARD"
        summary['A3'] = "KEY METRICS"
        summary['A4'] = "Total Revenue"
        summary['B4'] = "=Data!B15"
        summary['A5'] = "Total Customers"
        summary['B5'] = "=Data!B16"
        summary['A6'] = "Average Order Value"
        summary['B6'] = "=Data!B17"
        summary['A7'] = "Customer Lifetime Value"
        summary['B7'] = "=B4/B5*12"  # Simple CLV calculation
        
        # Trends section
        summary['A9'] = "TRENDS"
        summary['A10'] = "Revenue Growth"
        summary['B10'] = "=Data!C15"
        summary['A11'] = "Customer Growth"
        summary['B11'] = "=Data!C16"
        
        # Data sheet
        data_sheet = wb.create_sheet("Data")
        
        # Raw data
        data_sheet['A1'] = "Metric"
        data_sheet['B1'] = "Current"
        data_sheet['C1'] = "Growth %"
        
        data_sheet['A15'] = "Revenue"
        data_sheet['B15'] = 5000000
        data_sheet['C15'] = 0.23  # 23% growth
        
        data_sheet['A16'] = "Customers"
        data_sheet['B16'] = 25000
        data_sheet['C16'] = 0.15  # 15% growth
        
        data_sheet['A17'] = "AOV"
        data_sheet['B17'] = "=B15/B16"
        
        # Extract and verify
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        result = self.excel_extractor.extract_from_bytes(excel_bytes.getvalue(), "dashboard.xlsx")
        
        # Verify dashboard structure
        dashboard_sheet = result['sheets']['Dashboard']
        self.assertTrue(len(dashboard_sheet['key_metrics']) > 0)
        
        # Verify cross-sheet formula references
        dashboard_formulas = dashboard_sheet['formulas']
        data_references = [f for f in dashboard_formulas.values() 
                          if 'Data!' in f['formula']]
        self.assertTrue(len(data_references) > 0)

    def test_source_attribution_accuracy(self):
        """Test accuracy of source attribution across complex workbook"""
        wb = self.create_complex_workbook()
        
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        # Extract with source tracking
        result = self.excel_extractor.extract_from_bytes(excel_bytes.getvalue(), "attribution_test.xlsx")
        
        # Verify attribution data
        self.assertIn('_attribution', result)
        attribution_data = result['_attribution']['source_tracker_data']
        
        # Verify data points have proper source attribution
        data_points = attribution_data['data_points']
        self.assertTrue(len(data_points) > 0)
        
        for dp_id, dp_data in data_points.items():
            # Verify each data point has complete source information
            source = dp_data['primary_source']
            self.assertIn('document_name', source)
            self.assertIn('document_type', source)
            self.assertIn('page_or_sheet', source)
            self.assertIn('cell_or_section', source)
            self.assertEqual(source['document_type'], 'excel')
            self.assertEqual(source['document_name'], 'attribution_test.xlsx')

    def create_complex_workbook(self):
        """Create a complex workbook for comprehensive testing"""
        wb = Workbook()
        
        # Sheet 1: Financial Data
        fin_sheet = wb.active
        fin_sheet.title = "Financials"
        
        # Multi-year financial data
        years = [2021, 2022, 2023]
        fin_sheet['A1'] = "Metric"
        for i, year in enumerate(years):
            fin_sheet.cell(row=1, column=i+2, value=year)
        
        # Revenue data
        fin_sheet['A2'] = "Revenue"
        revenues = [8000000, 9500000, 12000000]
        for i, revenue in enumerate(revenues):
            fin_sheet.cell(row=2, column=i+2, value=revenue)
        
        # Growth calculations
        fin_sheet['A3'] = "Revenue Growth"
        fin_sheet['B3'] = None  # No growth for first year
        fin_sheet['C3'] = "=(C2-B2)/B2"  # 2022 growth
        fin_sheet['D3'] = "=(D2-C2)/C2"  # 2023 growth
        
        # Sheet 2: Operational Metrics
        ops_sheet = wb.create_sheet("Operations")
        ops_sheet['A1'] = "Customer Metrics"
        ops_sheet['A2'] = "Total Customers"
        ops_sheet['B2'] = 50000
        ops_sheet['A3'] = "New Customers (Monthly)"
        ops_sheet['B3'] = 2500
        ops_sheet['A4'] = "Churn Rate"
        ops_sheet['B4'] = 0.05  # 5%
        ops_sheet['A5'] = "Net Customer Growth"
        ops_sheet['B5'] = "=B3*(1-B4)"
        
        return wb


if __name__ == '__main__':
    unittest.main()