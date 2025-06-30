"""
Comprehensive tests for source attribution and data point tracking accuracy
Tests validate that source tracking works correctly across all extractors
"""

import unittest
from unittest.mock import Mock, patch
import sys
import os
import io
import json
from openpyxl import Workbook
from docx import Document

# Add the lib directory to the path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..', 'lib'))

from pdf_extractor import PDFExtractor
from excel_extractor import ExcelExtractor
from word_extractor import WordExtractor
from source_tracker import SourceTracker


class TestSourceAttribution(unittest.TestCase):
    """Test suite for source attribution accuracy"""

    def setUp(self):
        """Set up test fixtures"""
        self.source_tracker = SourceTracker()
        self.pdf_extractor = PDFExtractor(source_tracker=self.source_tracker)
        self.excel_extractor = ExcelExtractor(source_tracker=self.source_tracker)
        
    def test_source_tracker_basic_functionality(self):
        """Test basic source tracker functionality"""
        # Register a document
        doc_id = self.source_tracker.register_document("test.xlsx", "excel", {"test": "metadata"})
        
        self.assertIsNotNone(doc_id)
        self.assertIn(doc_id, self.source_tracker.documents)
        self.assertEqual(self.source_tracker.documents[doc_id]['type'], 'excel')
        
        # Track a data point
        data_point_id = self.source_tracker.track_data_point(
            value=1000000,
            document_id=doc_id,
            location_details={
                'page_or_sheet': 'Sheet1',
                'cell_or_section': 'B2',
                'coordinates': {'row': 2, 'col': 2}
            },
            confidence=0.95,
            context="Revenue figure"
        )
        
        self.assertIsNotNone(data_point_id)
        self.assertIn(data_point_id, self.source_tracker.data_points)
        
        # Verify data point details
        data_point = self.source_tracker.data_points[data_point_id]
        self.assertEqual(data_point.value, 1000000)
        self.assertEqual(data_point.confidence, 0.95)
        self.assertEqual(data_point.primary_source.document_id, doc_id)
        self.assertEqual(data_point.primary_source.cell_or_section, 'B2')

    def test_excel_source_attribution_accuracy(self):
        """Test accuracy of Excel source attribution"""
        # Create test workbook with known data
        wb = Workbook()
        ws = wb.active
        ws.title = "Financial Data"
        
        # Add specific test data with known locations
        ws['A1'] = "Metric"
        ws['B1'] = "Value"
        ws['A2'] = "Revenue"
        ws['B2'] = 1500000  # Target value at B2
        ws['A3'] = "Profit"
        ws['B3'] = "=B2*0.2"  # Formula at B3
        
        # Add another sheet
        ws2 = wb.create_sheet("Summary")
        ws2['A1'] = "Total"
        ws2['A2'] = "='Financial Data'!B2"  # Cross-sheet reference
        
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        result = self.excel_extractor.extract_from_bytes(excel_bytes.getvalue(), "financial.xlsx")
        
        # Verify attribution data was created
        self.assertIn('_attribution', result)
        attribution_data = result['_attribution']['source_tracker_data']
        
        # Check that data points were tracked
        data_points = attribution_data['data_points']
        self.assertTrue(len(data_points) > 0)
        
        # Find the specific value we're tracking (1500000)
        target_data_point = None
        for dp_id, dp_data in data_points.items():
            if dp_data['value'] == 1500000:
                target_data_point = dp_data
                break
        
        self.assertIsNotNone(target_data_point, "Revenue value should be tracked")
        
        # Verify attribution accuracy
        source = target_data_point['primary_source']
        self.assertEqual(source['document_name'], 'financial.xlsx')
        self.assertEqual(source['document_type'], 'excel')
        self.assertEqual(source['page_or_sheet'], 'Financial Data')
        self.assertEqual(source['cell_or_section'], 'B2')
        
        # Verify coordinates
        coords = source['coordinates']
        self.assertEqual(coords['row'], 2)
        self.assertEqual(coords['col'], 2)

    def test_excel_formula_attribution(self):
        """Test attribution of formula-based values"""
        wb = Workbook()
        ws = wb.active
        
        # Create data with formulas
        ws['A1'] = 1000
        ws['A2'] = 500
        ws['A3'] = "=A1+A2"  # Sum formula
        ws['A4'] = "=A3*0.1"  # Percentage calculation
        
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        result = self.excel_extractor.extract_from_bytes(excel_bytes.getvalue(), "formulas.xlsx")
        
        # Check formula tracking
        sheet_data = result['sheets']['Sheet']
        self.assertIn('formulas', sheet_data)
        
        # Verify A3 formula is tracked
        self.assertIn('A3', sheet_data['formulas'])
        formula_data = sheet_data['formulas']['A3']
        self.assertEqual(formula_data['formula'], '=A1+A2')
        
        # Verify formula has data point ID
        self.assertIsNotNone(formula_data.get('data_point_id'))
        
        # Check that calculated flag is set for formula-based data points
        attribution_data = result['_attribution']['source_tracker_data']
        formula_data_points = [
            dp for dp in attribution_data['data_points'].values()
            if dp.get('calculated') == True
        ]
        self.assertTrue(len(formula_data_points) > 0)

    def test_pdf_source_attribution_with_llmwhisperer(self):
        """Test PDF source attribution with mock LLMWhisperer responses"""
        with patch.dict(os.environ, {'LLMWHISPERER_API_KEY': 'test_key'}):
            with patch('pdf_extractor.requests.post') as mock_post, \
                 patch('pdf_extractor.requests.get') as mock_get:
                
                # Mock successful API responses
                mock_post.return_value.status_code = 200
                mock_post.return_value.json.return_value = {"whisper_hash": "test_hash"}
                
                # Mock extraction result with structured content
                extracted_text = """# Financial Report 2023
                
Revenue Performance:
- Total Revenue: $2,500,000
- Growth Rate: 35%

| Quarter | Revenue | Growth |
|---------|---------|--------|
| Q1 | $600,000 | 10% |
| Q2 | $650,000 | 15% |
| Q3 | $700,000 | 20% |
| Q4 | $550,000 | 5% |

Key Metrics:
- EBITDA: $750,000
- Net Profit: $450,000
"""
                
                mock_get.side_effect = [
                    Mock(status_code=200, json=lambda: {"status": "processed"}),
                    Mock(status_code=200, json=lambda: {"extracted_text": extracted_text})
                ]
                
                result = self.pdf_extractor.extract_from_bytes(b"fake_pdf", "report.pdf")
                
                # Verify attribution data
                self.assertIn('_attribution', result)
                attribution_data = result['_attribution']['source_tracker_data']
                
                # Check that financial values were tracked
                data_points = attribution_data['data_points']
                financial_data_points = [
                    dp for dp in data_points.values()
                    if dp['data_type'] in ['financial', 'financial_large', 'financial_medium']
                ]
                self.assertTrue(len(financial_data_points) > 0)
                
                # Verify page number attribution
                for dp_data in financial_data_points:
                    source = dp_data['primary_source']
                    self.assertEqual(source['document_name'], 'report.pdf')
                    self.assertEqual(source['document_type'], 'pdf')
                    self.assertIn('Page', source['page_or_sheet'])

    def test_cross_reference_validation(self):
        """Test validation of cross-referenced data points"""
        # Create Excel with cross-sheet references
        wb = Workbook()
        
        # Sheet 1: Raw data
        data_sheet = wb.active
        data_sheet.title = "Data"
        data_sheet['A1'] = "Revenue"
        data_sheet['B1'] = 1000000
        
        # Sheet 2: Summary with reference
        summary_sheet = wb.create_sheet("Summary")
        summary_sheet['A1'] = "Total Revenue"
        summary_sheet['B1'] = "=Data!B1"
        
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        result = self.excel_extractor.extract_from_bytes(excel_bytes.getvalue(), "cross_ref.xlsx")
        
        # Get all data point IDs
        attribution_data = result['_attribution']['source_tracker_data']
        data_point_ids = list(attribution_data['data_points'].keys())
        
        # Test validation across data points
        validation = self.source_tracker.validate_data_consistency(data_point_ids)
        
        self.assertTrue(validation['consistent'])
        self.assertEqual(validation['source_coverage']['unique_documents'], 1)
        self.assertGreaterEqual(validation['confidence_distribution']['average'], 0.8)

    def test_source_hyperlink_generation(self):
        """Test generation of source hyperlinks for PowerPoint"""
        # Register document and track data point
        doc_id = self.source_tracker.register_document("test.xlsx", "excel")
        
        data_point_id = self.source_tracker.track_data_point(
            value=1500000,
            document_id=doc_id,
            location_details={
                'page_or_sheet': 'Financial Data',
                'cell_or_section': 'B15'
            }
        )
        
        # Test hyperlink generation
        hyperlink = self.source_tracker.get_source_hyperlink(data_point_id)
        
        self.assertIn('file:///', hyperlink)
        self.assertIn('test.xlsx', hyperlink)
        self.assertIn('Financial Data', hyperlink)
        self.assertIn('B15', hyperlink)
        
        # Test with custom link text
        custom_hyperlink = self.source_tracker.get_source_hyperlink(data_point_id, "Custom Link")
        self.assertIn('file:///', custom_hyperlink)

    def test_source_attribution_text_formatting(self):
        """Test various formats of source attribution text"""
        doc_id = self.source_tracker.register_document("financial_report.xlsx", "excel")
        
        data_point_id = self.source_tracker.track_data_point(
            value=2500000,
            document_id=doc_id,
            location_details={
                'page_or_sheet': 'Income Statement',
                'cell_or_section': 'C10'
            },
            confidence=0.95
        )
        
        # Test minimal format
        minimal = self.source_tracker.get_source_attribution_text(data_point_id, 'minimal')
        self.assertEqual(minimal, "Source: financial_report.xlsx")
        
        # Test detailed format
        detailed = self.source_tracker.get_source_attribution_text(data_point_id, 'detailed')
        self.assertIn("Source: financial_report.xlsx")
        self.assertIn("Sheet: Income Statement")
        self.assertIn("Location: C10")
        self.assertIn("Confidence: 95.0%")
        
        # Test comprehensive format with secondary sources
        self.source_tracker.add_secondary_source(
            data_point_id, doc_id,
            {'page_or_sheet': 'Summary', 'cell_or_section': 'A5'}
        )
        
        comprehensive = self.source_tracker.get_source_attribution_text(data_point_id, 'comprehensive')
        self.assertIn("Also in:")
        self.assertIn("financial_report.xlsx:A5")

    def test_confidence_scoring_accuracy(self):
        """Test accuracy of confidence scoring for different scenarios"""
        wb = Workbook()
        ws = wb.active
        
        # High confidence scenario: Formula with good context
        ws['A1'] = "Total Revenue"
        ws['B1'] = "=SUM(B2:B5)"
        ws['B2'] = 250000
        ws['B3'] = 300000
        ws['B4'] = 275000
        ws['B5'] = 325000
        
        # Medium confidence scenario: Large number without clear context
        ws['C1'] = 1500000
        
        # Lower confidence scenario: Text that might be data
        ws['D1'] = "Some text value"
        
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        result = self.excel_extractor.extract_from_bytes(excel_bytes.getvalue(), "confidence.xlsx")
        
        # Analyze confidence scores
        attribution_data = result['_attribution']['source_tracker_data']
        data_points = attribution_data['data_points']
        
        confidences = [dp['confidence'] for dp in data_points.values()]
        
        # Should have a range of confidence scores
        self.assertGreater(max(confidences), 0.9)  # At least one high confidence
        self.assertLess(min(confidences), 0.9)     # At least one lower confidence

    def test_context_preservation_accuracy(self):
        """Test that context is accurately preserved and retrievable"""
        doc_id = self.source_tracker.register_document("context_test.xlsx", "excel")
        
        # Track data point with rich context
        data_point_id = self.source_tracker.track_data_point(
            value=1750000,
            document_id=doc_id,
            location_details={
                'page_or_sheet': 'Financial Analysis',
                'cell_or_section': 'D15',
                'table_name': 'Annual Revenue Summary',
                'coordinates': {'row': 15, 'col': 4},
                'extraction_method': 'openpyxl'
            },
            confidence=0.95,
            context="Revenue figure in annual summary table, surrounded by quarterly breakdown",
            formula="=SUM(D2:D14)"
        )
        
        # Retrieve and verify context
        context = self.source_tracker.get_source_context(data_point_id)
        
        self.assertIn('data_point', context)
        self.assertIn('quality_assessment', context)
        self.assertIn('source_details', context)
        self.assertIn('validation', context)
        
        # Verify specific context details
        source_details = context['source_details']
        self.assertEqual(source_details['document'], 'context_test.xlsx')
        self.assertEqual(source_details['type'], 'excel')
        self.assertEqual(source_details['location'], 'D15')
        
        quality = context['quality_assessment']
        self.assertEqual(quality['confidence'], 0.95)
        self.assertTrue(quality['has_formula'])
        self.assertTrue(quality['calculated'])

    def test_data_point_export_import_accuracy(self):
        """Test accurate export and import of attribution data"""
        # Create and populate source tracker
        doc_id = self.source_tracker.register_document("export_test.xlsx", "excel")
        
        data_point_ids = []
        for i in range(5):
            dp_id = self.source_tracker.track_data_point(
                value=1000 * (i + 1),
                document_id=doc_id,
                location_details={
                    'page_or_sheet': 'Sheet1',
                    'cell_or_section': f'A{i+1}',
                    'coordinates': {'row': i+1, 'col': 1}
                },
                confidence=0.9
            )
            data_point_ids.append(dp_id)
        
        # Export data
        exported_data = self.source_tracker.export_attribution_data()
        
        # Verify export structure
        self.assertIn('data_points', exported_data)
        self.assertIn('documents', exported_data)
        self.assertIn('source_mappings', exported_data)
        self.assertIn('metadata', exported_data)
        
        # Create new tracker and import
        new_tracker = SourceTracker()
        new_tracker.import_attribution_data(exported_data)
        
        # Verify imported data
        self.assertEqual(len(new_tracker.data_points), 5)
        self.assertEqual(len(new_tracker.documents), 1)
        
        # Verify specific data point accuracy
        for dp_id in data_point_ids:
            self.assertIn(dp_id, new_tracker.data_points)
            original_dp = self.source_tracker.data_points[dp_id]
            imported_dp = new_tracker.data_points[dp_id]
            
            self.assertEqual(original_dp.value, imported_dp.value)
            self.assertEqual(original_dp.confidence, imported_dp.confidence)
            self.assertEqual(original_dp.primary_source.cell_or_section, 
                           imported_dp.primary_source.cell_or_section)

    def test_multi_document_attribution_accuracy(self):
        """Test source attribution accuracy across multiple documents"""
        # Create multiple documents
        doc1_id = self.source_tracker.register_document("financial.xlsx", "excel")
        doc2_id = self.source_tracker.register_document("operational.pdf", "pdf")
        
        # Track data points from both documents
        excel_dp = self.source_tracker.track_data_point(
            value=1500000,
            document_id=doc1_id,
            location_details={'page_or_sheet': 'Sheet1', 'cell_or_section': 'B2'}
        )
        
        pdf_dp = self.source_tracker.track_data_point(
            value="25%",
            document_id=doc2_id,
            location_details={'page_or_sheet': 'Page 3', 'cell_or_section': 'Line 45'}
        )
        
        # Verify documents are tracked separately
        self.assertNotEqual(excel_dp, pdf_dp)
        
        excel_source = self.source_tracker.data_points[excel_dp].primary_source
        pdf_source = self.source_tracker.data_points[pdf_dp].primary_source
        
        self.assertEqual(excel_source.document_type, 'excel')
        self.assertEqual(pdf_source.document_type, 'pdf')
        self.assertNotEqual(excel_source.document_id, pdf_source.document_id)
        
        # Test validation across both documents
        validation = self.source_tracker.validate_data_consistency([excel_dp, pdf_dp])
        self.assertEqual(validation['source_coverage']['unique_documents'], 2)

    def test_table_cell_attribution_precision(self):
        """Test precise attribution of table cells"""
        # Test PDF table attribution
        table_lines = [
            "| Product | Q1 Revenue | Q2 Revenue |",
            "|---------|------------|------------|",
            "| Product A | $250,000 | $275,000 |",
            "| Product B | $300,000 | $325,000 |"
        ]
        
        doc_id = self.source_tracker.register_document("table_test.pdf", "pdf")
        
        data_point_ids = self.pdf_extractor._track_table_cells(table_lines, doc_id, 10)
        
        # Verify precise cell tracking
        self.assertTrue(len(data_point_ids) > 0)
        
        # Check that tracked data points have correct table context
        for dp_id in data_point_ids:
            if dp_id in self.source_tracker.data_points:
                dp = self.source_tracker.data_points[dp_id]
                location = dp.primary_source
                
                # Should have table-specific location details
                self.assertIn('Table', location.cell_or_section)
                self.assertEqual(location.coordinates['table_start'], 10)
                self.assertIn(location.coordinates['header'], ['Product', 'Q1 Revenue', 'Q2 Revenue'])


if __name__ == '__main__':
    unittest.main()