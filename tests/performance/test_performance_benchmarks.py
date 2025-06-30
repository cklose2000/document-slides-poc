#!/usr/bin/env python3
"""
Performance benchmarks for document processing pipeline
Tests processing speed, memory usage, and scalability
"""
import time
import psutil
import os
import sys
import tempfile
import io
import json
import statistics
from typing import Dict, List, Tuple, Any
from contextlib import contextmanager
from dataclasses import dataclass, asdict
import unittest
import concurrent.futures
import threading

# Add parent directories to path
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from lib.excel_extractor import ExcelExtractor
from lib.word_extractor import WordExtractor
from lib.pdf_extractor import PDFExtractor
from lib.slide_generator import SlideGenerator
from lib.llm_slides import analyze_documents_for_slides

@dataclass
class PerformanceMetrics:
    """Performance metrics data structure"""
    test_name: str
    execution_time: float
    memory_usage_mb: float
    peak_memory_mb: float
    cpu_usage_percent: float
    file_size_mb: float
    throughput_mb_per_sec: float
    success: bool
    error_message: str = ""
    
    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)

class PerformanceMonitor:
    """Context manager for monitoring performance metrics"""
    
    def __init__(self, test_name: str):
        self.test_name = test_name
        self.start_time = 0
        self.end_time = 0
        self.start_memory = 0
        self.peak_memory = 0
        self.cpu_samples = []
        self.monitoring = False
        self.monitor_thread = None
        
    def __enter__(self):
        self.start_time = time.time()
        process = psutil.Process()
        self.start_memory = process.memory_info().rss / 1024 / 1024  # MB
        self.peak_memory = self.start_memory
        self.monitoring = True
        
        # Start CPU monitoring thread
        self.monitor_thread = threading.Thread(target=self._monitor_resources)
        self.monitor_thread.daemon = True
        self.monitor_thread.start()
        
        return self
        
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.end_time = time.time()
        self.monitoring = False
        if self.monitor_thread:
            self.monitor_thread.join(timeout=1)
            
    def _monitor_resources(self):
        """Monitor CPU and memory usage in background thread"""
        process = psutil.Process()
        while self.monitoring:
            try:
                # Sample CPU usage
                cpu_percent = process.cpu_percent()
                self.cpu_samples.append(cpu_percent)
                
                # Track peak memory
                current_memory = process.memory_info().rss / 1024 / 1024
                self.peak_memory = max(self.peak_memory, current_memory)
                
                time.sleep(0.1)  # Sample every 100ms
            except:
                break
                
    def get_metrics(self, file_size_mb: float = 0, success: bool = True, error_message: str = "") -> PerformanceMetrics:
        """Calculate and return performance metrics"""
        execution_time = self.end_time - self.start_time
        avg_cpu = statistics.mean(self.cpu_samples) if self.cpu_samples else 0
        
        process = psutil.Process()
        final_memory = process.memory_info().rss / 1024 / 1024
        memory_usage = final_memory - self.start_memory
        
        throughput = file_size_mb / execution_time if execution_time > 0 and file_size_mb > 0 else 0
        
        return PerformanceMetrics(
            test_name=self.test_name,
            execution_time=execution_time,
            memory_usage_mb=memory_usage,
            peak_memory_mb=self.peak_memory,
            cpu_usage_percent=avg_cpu,
            file_size_mb=file_size_mb,
            throughput_mb_per_sec=throughput,
            success=success,
            error_message=error_message
        )

class DocumentProcessingBenchmarks(unittest.TestCase):
    """Performance benchmarks for document processing components"""
    
    @classmethod
    def setUpClass(cls):
        """Set up test environment"""
        cls.test_results = []
        cls.extractors = {
            'excel': ExcelExtractor(),
            'word': WordExtractor(),
            'pdf': PDFExtractor()
        }
        
        # Create test data directory
        cls.test_data_dir = os.path.join(os.path.dirname(__file__), 'test_data')
        os.makedirs(cls.test_data_dir, exist_ok=True)
        
    def setUp(self):
        """Set up each test"""
        self.temp_files = []
        
    def tearDown(self):
        """Clean up temporary files"""
        for temp_file in self.temp_files:
            try:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
            except:
                pass
                
    def create_test_excel(self, num_rows: int = 1000, num_cols: int = 10) -> str:
        """Create test Excel file with specified dimensions"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Add headers
        for col in range(1, num_cols + 1):
            ws.cell(row=1, column=col, value=f"Column_{col}")
            
        # Add data rows
        for row in range(2, num_rows + 2):
            for col in range(1, num_cols + 1):
                ws.cell(row=row, column=col, value=f"Data_{row}_{col}")
                
        temp_path = tempfile.mktemp(suffix='.xlsx')
        wb.save(temp_path)
        self.temp_files.append(temp_path)
        return temp_path
        
    def create_test_word_doc(self, num_paragraphs: int = 100) -> str:
        """Create test Word document with specified content"""
        from docx import Document
        doc = Document()
        
        for i in range(num_paragraphs):
            doc.add_paragraph(f"This is paragraph {i + 1}. " * 10)
            
        temp_path = tempfile.mktemp(suffix='.docx')
        doc.save(temp_path)
        self.temp_files.append(temp_path)
        return temp_path
        
    def get_file_size_mb(self, file_path: str) -> float:
        """Get file size in MB"""
        return os.path.getsize(file_path) / 1024 / 1024
        
    def test_excel_extraction_small(self):
        """Benchmark Excel extraction - small file"""
        file_path = self.create_test_excel(100, 5)
        file_size = self.get_file_size_mb(file_path)
        
        with PerformanceMonitor("excel_extraction_small") as monitor:
            try:
                result = self.extractors['excel'].extract_from_file(file_path)
                success = len(result) > 0
            except Exception as e:
                success = False
                error_msg = str(e)
            else:
                error_msg = ""
                
        metrics = monitor.get_metrics(file_size, success, error_msg)
        self.test_results.append(metrics)
        self.assertTrue(success, f"Excel extraction failed: {error_msg}")
        
    def test_excel_extraction_medium(self):
        """Benchmark Excel extraction - medium file"""
        file_path = self.create_test_excel(1000, 20)
        file_size = self.get_file_size_mb(file_path)
        
        with PerformanceMonitor("excel_extraction_medium") as monitor:
            try:
                result = self.extractors['excel'].extract_from_file(file_path)
                success = len(result) > 0
            except Exception as e:
                success = False
                error_msg = str(e)
            else:
                error_msg = ""
                
        metrics = monitor.get_metrics(file_size, success, error_msg)
        self.test_results.append(metrics)
        self.assertTrue(success, f"Excel extraction failed: {error_msg}")
        
    def test_excel_extraction_large(self):
        """Benchmark Excel extraction - large file"""
        file_path = self.create_test_excel(5000, 50)
        file_size = self.get_file_size_mb(file_path)
        
        with PerformanceMonitor("excel_extraction_large") as monitor:
            try:
                result = self.extractors['excel'].extract_from_file(file_path)
                success = len(result) > 0
            except Exception as e:
                success = False
                error_msg = str(e)
            else:
                error_msg = ""
                
        metrics = monitor.get_metrics(file_size, success, error_msg)
        self.test_results.append(metrics)
        self.assertTrue(success, f"Excel extraction failed: {error_msg}")
        
    def test_word_extraction_small(self):
        """Benchmark Word extraction - small document"""
        file_path = self.create_test_word_doc(50)
        file_size = self.get_file_size_mb(file_path)
        
        with PerformanceMonitor("word_extraction_small") as monitor:
            try:
                result = self.extractors['word'].extract_from_file(file_path)
                success = len(result) > 0
            except Exception as e:
                success = False
                error_msg = str(e)
            else:
                error_msg = ""
                
        metrics = monitor.get_metrics(file_size, success, error_msg)
        self.test_results.append(metrics)
        self.assertTrue(success, f"Word extraction failed: {error_msg}")
        
    def test_word_extraction_large(self):
        """Benchmark Word extraction - large document"""
        file_path = self.create_test_word_doc(500)
        file_size = self.get_file_size_mb(file_path)
        
        with PerformanceMonitor("word_extraction_large") as monitor:
            try:
                result = self.extractors['word'].extract_from_file(file_path)
                success = len(result) > 0
            except Exception as e:
                success = False
                error_msg = str(e)
            else:
                error_msg = ""
                
        metrics = monitor.get_metrics(file_size, success, error_msg)
        self.test_results.append(metrics)
        self.assertTrue(success, f"Word extraction failed: {error_msg}")
        
    def test_slide_generation_performance(self):
        """Benchmark slide generation with various content sizes"""
        content_sizes = [
            ("small", ["Short content"] * 5),
            ("medium", ["Medium length content with more details"] * 15),
            ("large", ["Long content with extensive details and comprehensive information"] * 30)
        ]
        
        for size_name, content in content_sizes:
            with PerformanceMonitor(f"slide_generation_{size_name}") as monitor:
                try:
                    generator = SlideGenerator()
                    
                    # Create slides with content
                    for i, text in enumerate(content):
                        generator.add_title_slide(f"Slide {i + 1}", text)
                        
                    # Save to temporary file
                    temp_path = tempfile.mktemp(suffix='.pptx')
                    self.temp_files.append(temp_path)
                    generator.save(temp_path)
                    
                    success = os.path.exists(temp_path)
                    file_size = self.get_file_size_mb(temp_path) if success else 0
                except Exception as e:
                    success = False
                    error_msg = str(e)
                    file_size = 0
                else:
                    error_msg = ""
                    
            metrics = monitor.get_metrics(file_size, success, error_msg)
            self.test_results.append(metrics)
            self.assertTrue(success, f"Slide generation failed: {error_msg}")
            
    def test_concurrent_processing(self):
        """Benchmark concurrent document processing"""
        num_concurrent = 5
        
        # Create multiple test files
        test_files = []
        for i in range(num_concurrent):
            file_path = self.create_test_excel(200, 10)
            test_files.append(file_path)
            
        with PerformanceMonitor("concurrent_processing") as monitor:
            try:
                with concurrent.futures.ThreadPoolExecutor(max_workers=num_concurrent) as executor:
                    futures = []
                    for file_path in test_files:
                        future = executor.submit(self.extractors['excel'].extract_from_file, file_path)
                        futures.append(future)
                        
                    results = []
                    for future in concurrent.futures.as_completed(futures):
                        result = future.result()
                        results.append(result)
                        
                success = len(results) == num_concurrent
                total_size = sum(self.get_file_size_mb(f) for f in test_files)
            except Exception as e:
                success = False
                error_msg = str(e)
                total_size = 0
            else:
                error_msg = ""
                
        metrics = monitor.get_metrics(total_size, success, error_msg)
        self.test_results.append(metrics)
        self.assertTrue(success, f"Concurrent processing failed: {error_msg}")
        
    @classmethod
    def tearDownClass(cls):
        """Generate performance report"""
        cls.generate_performance_report()
        
    @classmethod
    def generate_performance_report(cls):
        """Generate comprehensive performance report"""
        if not cls.test_results:
            return
            
        report_data = {
            'test_suite': 'Document Processing Performance Benchmarks',
            'timestamp': time.strftime('%Y-%m-%d %H:%M:%S'),
            'total_tests': len(cls.test_results),
            'successful_tests': sum(1 for r in cls.test_results if r.success),
            'failed_tests': sum(1 for r in cls.test_results if not r.success),
            'metrics': [r.to_dict() for r in cls.test_results],
            'summary': cls._calculate_summary_statistics()
        }
        
        # Save JSON report
        report_path = f"performance_report_{int(time.time())}.json"
        with open(report_path, 'w') as f:
            json.dump(report_data, f, indent=2)
            
        # Generate readable report
        readable_path = f"performance_report_{int(time.time())}.txt"
        with open(readable_path, 'w') as f:
            cls._write_readable_report(f, report_data)
            
        print(f"\nPerformance reports generated:")
        print(f"  JSON: {os.path.abspath(report_path)}")
        print(f"  Text: {os.path.abspath(readable_path)}")
        
    @classmethod
    def _calculate_summary_statistics(cls) -> Dict[str, Any]:
        """Calculate summary statistics from test results"""
        successful_results = [r for r in cls.test_results if r.success]
        
        if not successful_results:
            return {}
            
        execution_times = [r.execution_time for r in successful_results]
        memory_usage = [r.memory_usage_mb for r in successful_results]
        throughput = [r.throughput_mb_per_sec for r in successful_results if r.throughput_mb_per_sec > 0]
        
        return {
            'avg_execution_time': statistics.mean(execution_times),
            'max_execution_time': max(execution_times),
            'min_execution_time': min(execution_times),
            'avg_memory_usage_mb': statistics.mean(memory_usage),
            'max_memory_usage_mb': max(memory_usage),
            'avg_throughput_mb_per_sec': statistics.mean(throughput) if throughput else 0,
            'max_throughput_mb_per_sec': max(throughput) if throughput else 0
        }
        
    @classmethod
    def _write_readable_report(cls, f, report_data):
        """Write human-readable performance report"""
        f.write("DOCUMENT PROCESSING PERFORMANCE BENCHMARK REPORT\n")
        f.write("=" * 60 + "\n\n")
        
        f.write(f"Test Suite: {report_data['test_suite']}\n")
        f.write(f"Timestamp: {report_data['timestamp']}\n")
        f.write(f"Total Tests: {report_data['total_tests']}\n")
        f.write(f"Successful: {report_data['successful_tests']}\n")
        f.write(f"Failed: {report_data['failed_tests']}\n\n")
        
        # Summary statistics
        if report_data['summary']:
            f.write("SUMMARY STATISTICS\n")
            f.write("-" * 30 + "\n")
            summary = report_data['summary']
            f.write(f"Average Execution Time: {summary.get('avg_execution_time', 0):.3f} seconds\n")
            f.write(f"Max Execution Time: {summary.get('max_execution_time', 0):.3f} seconds\n")
            f.write(f"Average Memory Usage: {summary.get('avg_memory_usage_mb', 0):.2f} MB\n")
            f.write(f"Max Memory Usage: {summary.get('max_memory_usage_mb', 0):.2f} MB\n")
            f.write(f"Average Throughput: {summary.get('avg_throughput_mb_per_sec', 0):.2f} MB/s\n")
            f.write(f"Max Throughput: {summary.get('max_throughput_mb_per_sec', 0):.2f} MB/s\n\n")
        
        # Individual test results
        f.write("INDIVIDUAL TEST RESULTS\n")
        f.write("-" * 30 + "\n")
        
        for result in report_data['metrics']:
            f.write(f"\nTest: {result['test_name']}\n")
            f.write(f"  Status: {'PASS' if result['success'] else 'FAIL'}\n")
            if not result['success']:
                f.write(f"  Error: {result['error_message']}\n")
            f.write(f"  Execution Time: {result['execution_time']:.3f} seconds\n")
            f.write(f"  Memory Usage: {result['memory_usage_mb']:.2f} MB\n")
            f.write(f"  Peak Memory: {result['peak_memory_mb']:.2f} MB\n")
            f.write(f"  CPU Usage: {result['cpu_usage_percent']:.1f}%\n")
            f.write(f"  File Size: {result['file_size_mb']:.2f} MB\n")
            f.write(f"  Throughput: {result['throughput_mb_per_sec']:.2f} MB/s\n")

if __name__ == '__main__':
    # Create test directory
    os.makedirs('tests/performance', exist_ok=True)
    
    # Run performance benchmarks
    unittest.main(verbosity=2)