#!/usr/bin/env python3
"""
Large file handling and memory usage stress tests
Tests system behavior with large files, memory limits, and resource constraints
"""
import os
import sys
import tempfile
import io
import time
import unittest
import psutil
import threading
import gc
from typing import Dict, List, Tuple, Any, Optional
from dataclasses import dataclass
import json
import statistics

# Add parent directories to path
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from lib.excel_extractor import ExcelExtractor
from lib.word_extractor import WordExtractor
from lib.pdf_extractor import PDFExtractor
from lib.slide_generator import SlideGenerator

@dataclass
class MemoryProfile:
    """Memory usage profile data"""
    test_name: str
    initial_memory_mb: float
    peak_memory_mb: float
    final_memory_mb: float
    memory_delta_mb: float
    peak_delta_mb: float
    execution_time_sec: float
    gc_collections: int
    success: bool
    error_message: str = ""

class MemoryProfiler:
    """Context manager for memory profiling"""
    
    def __init__(self, test_name: str):
        self.test_name = test_name
        self.initial_memory = 0
        self.peak_memory = 0
        self.final_memory = 0
        self.start_time = 0
        self.end_time = 0
        self.initial_gc_count = 0
        self.monitoring = False
        self.monitor_thread = None
        
    def __enter__(self):
        # Force garbage collection before starting
        gc.collect()
        
        process = psutil.Process()
        self.initial_memory = process.memory_info().rss / 1024 / 1024  # MB
        self.peak_memory = self.initial_memory
        self.start_time = time.time()
        self.initial_gc_count = sum(gc.get_count())
        
        # Start monitoring thread
        self.monitoring = True
        self.monitor_thread = threading.Thread(target=self._monitor_memory)
        self.monitor_thread.daemon = True
        self.monitor_thread.start()
        
        return self
        
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.end_time = time.time()
        self.monitoring = False
        
        if self.monitor_thread:
            self.monitor_thread.join(timeout=1)
            
        # Final memory measurement
        process = psutil.Process()
        self.final_memory = process.memory_info().rss / 1024 / 1024
        
        # Force garbage collection and measure again
        gc.collect()
        final_after_gc = process.memory_info().rss / 1024 / 1024
        
        return False  # Don't suppress exceptions
        
    def _monitor_memory(self):
        """Monitor memory usage in background thread"""
        process = psutil.Process()
        while self.monitoring:
            try:
                current_memory = process.memory_info().rss / 1024 / 1024
                self.peak_memory = max(self.peak_memory, current_memory)
                time.sleep(0.1)  # Sample every 100ms
            except:
                break
                
    def get_profile(self, success: bool = True, error_message: str = "") -> MemoryProfile:
        """Get memory profile results"""
        final_gc_count = sum(gc.get_count())
        
        return MemoryProfile(
            test_name=self.test_name,
            initial_memory_mb=self.initial_memory,
            peak_memory_mb=self.peak_memory,
            final_memory_mb=self.final_memory,
            memory_delta_mb=self.final_memory - self.initial_memory,
            peak_delta_mb=self.peak_memory - self.initial_memory,
            execution_time_sec=self.end_time - self.start_time,
            gc_collections=final_gc_count - self.initial_gc_count,
            success=success,
            error_message=error_message
        )

class LargeFileStressTests(unittest.TestCase):
    """Stress tests for large file handling"""
    
    @classmethod
    def setUpClass(cls):
        """Set up test environment"""
        cls.memory_profiles = []
        cls.temp_files = []
        cls.extractors = {
            'excel': ExcelExtractor(),
            'word': WordExtractor(),
            'pdf': PDFExtractor()
        }
        
        # Create test data directory
        cls.test_data_dir = os.path.join(os.path.dirname(__file__), 'test_data')
        os.makedirs(cls.test_data_dir, exist_ok=True)
        
    @classmethod
    def tearDownClass(cls):
        """Clean up and generate report"""
        # Clean up temporary files
        for temp_file in cls.temp_files:
            try:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
            except:
                pass
                
        # Generate memory usage report
        cls.generate_memory_report()
        
    def setUp(self):
        """Set up each test"""
        # Check available memory before each test
        available_memory = psutil.virtual_memory().available / 1024 / 1024  # MB
        if available_memory < 1000:  # Less than 1GB available
            self.skipTest("Insufficient memory available for stress testing")
            
    def create_large_excel(self, rows: int, cols: int) -> str:
        """Create large Excel file with specified dimensions"""
        import openpyxl
        
        temp_path = tempfile.mktemp(suffix='.xlsx')
        self.temp_files.append(temp_path)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Add headers
        for col in range(1, cols + 1):
            ws.cell(row=1, column=col, value=f"Column_{col}")
            
        # Add data in chunks to manage memory
        chunk_size = 1000
        for start_row in range(2, rows + 2, chunk_size):
            end_row = min(start_row + chunk_size, rows + 2)
            
            for row in range(start_row, end_row):
                for col in range(1, cols + 1):
                    ws.cell(row=row, column=col, value=f"Data_{row}_{col}")
                    
            # Periodic garbage collection
            if start_row % (chunk_size * 10) == 0:
                gc.collect()
                
        wb.save(temp_path)
        return temp_path
        
    def create_large_word_doc(self, paragraphs: int, words_per_paragraph: int = 100) -> str:
        """Create large Word document"""
        from docx import Document
        
        temp_path = tempfile.mktemp(suffix='.docx')
        self.temp_files.append(temp_path)
        
        doc = Document()
        
        # Add paragraphs in chunks
        chunk_size = 100
        for start_para in range(0, paragraphs, chunk_size):
            end_para = min(start_para + chunk_size, paragraphs)
            
            for para_num in range(start_para, end_para):
                # Create paragraph with specified word count
                words = [f"word{i}" for i in range(words_per_paragraph)]
                paragraph_text = " ".join(words)
                doc.add_paragraph(f"Paragraph {para_num + 1}: {paragraph_text}")
                
            # Periodic garbage collection
            if start_para % (chunk_size * 10) == 0:
                gc.collect()
                
        doc.save(temp_path)
        return temp_path
        
    def get_file_size_mb(self, file_path: str) -> float:
        """Get file size in MB"""
        return os.path.getsize(file_path) / 1024 / 1024
        
    def test_large_excel_10mb(self):
        """Test processing ~10MB Excel file"""
        with MemoryProfiler("large_excel_10mb") as profiler:
            try:
                # Create ~10MB Excel file (approximately 50,000 rows x 20 columns)
                file_path = self.create_large_excel(50000, 20)
                file_size = self.get_file_size_mb(file_path)
                print(f"Created Excel file: {file_size:.2f} MB")
                
                # Process the file
                result = self.extractors['excel'].extract_from_file(file_path)
                
                success = result is not None
                error_msg = ""
                
            except Exception as e:
                success = False
                error_msg = str(e)
                
        profile = profiler.get_profile(success, error_msg)
        self.memory_profiles.append(profile)
        
        if success:
            print(f"Memory usage: {profile.peak_delta_mb:.2f} MB peak, {profile.memory_delta_mb:.2f} MB delta")
        else:
            print(f"Test failed: {error_msg}")
            
    def test_large_excel_50mb(self):
        """Test processing ~50MB Excel file"""
        with MemoryProfiler("large_excel_50mb") as profiler:
            try:
                # Create ~50MB Excel file (approximately 100,000 rows x 50 columns)
                file_path = self.create_large_excel(100000, 50)
                file_size = self.get_file_size_mb(file_path)
                print(f"Created Excel file: {file_size:.2f} MB")
                
                # Process the file
                result = self.extractors['excel'].extract_from_file(file_path)
                
                success = result is not None
                error_msg = ""
                
            except Exception as e:
                success = False
                error_msg = str(e)
                
        profile = profiler.get_profile(success, error_msg)
        self.memory_profiles.append(profile)
        
        if success:
            print(f"Memory usage: {profile.peak_delta_mb:.2f} MB peak, {profile.memory_delta_mb:.2f} MB delta")
        else:
            print(f"Test failed: {error_msg}")
            
    def test_large_word_10mb(self):
        """Test processing ~10MB Word document"""
        with MemoryProfiler("large_word_10mb") as profiler:
            try:
                # Create ~10MB Word document (approximately 5,000 paragraphs)
                file_path = self.create_large_word_doc(5000, 200)
                file_size = self.get_file_size_mb(file_path)
                print(f"Created Word document: {file_size:.2f} MB")
                
                # Process the file
                result = self.extractors['word'].extract_from_file(file_path)
                
                success = result is not None
                error_msg = ""
                
            except Exception as e:
                success = False
                error_msg = str(e)
                
        profile = profiler.get_profile(success, error_msg)
        self.memory_profiles.append(profile)
        
        if success:
            print(f"Memory usage: {profile.peak_delta_mb:.2f} MB peak, {profile.memory_delta_mb:.2f} MB delta")
        else:
            print(f"Test failed: {error_msg}")
            
    def test_large_slide_generation(self):
        """Test generating presentation with large amount of content"""
        with MemoryProfiler("large_slide_generation") as profiler:
            try:
                generator = SlideGenerator()
                
                # Generate 500 slides with substantial content
                num_slides = 500
                content_per_slide = "This is a large amount of content for testing memory usage. " * 50
                
                for i in range(num_slides):
                    generator.add_title_slide(f"Slide {i + 1}", content_per_slide)
                    
                    # Periodic progress updates
                    if (i + 1) % 100 == 0:
                        print(f"Generated {i + 1} slides...")
                        
                # Save the presentation
                temp_path = tempfile.mktemp(suffix='.pptx')
                self.temp_files.append(temp_path)
                generator.save(temp_path)
                
                file_size = self.get_file_size_mb(temp_path)
                print(f"Generated presentation: {file_size:.2f} MB")
                
                success = os.path.exists(temp_path)
                error_msg = ""
                
            except Exception as e:
                success = False
                error_msg = str(e)
                
        profile = profiler.get_profile(success, error_msg)
        self.memory_profiles.append(profile)
        
        if success:
            print(f"Memory usage: {profile.peak_delta_mb:.2f} MB peak, {profile.memory_delta_mb:.2f} MB delta")
        else:
            print(f"Test failed: {error_msg}")
            
    def test_memory_leak_detection(self):
        """Test for memory leaks in repeated operations"""
        with MemoryProfiler("memory_leak_detection") as profiler:
            try:
                initial_memory = psutil.Process().memory_info().rss / 1024 / 1024
                memory_samples = []
                
                # Perform repeated operations
                for iteration in range(20):
                    # Create and process small file
                    file_path = self.create_large_excel(1000, 10)
                    result = self.extractors['excel'].extract_from_file(file_path)
                    
                    # Clean up file immediately
                    os.remove(file_path)
                    
                    # Force garbage collection
                    gc.collect()
                    
                    # Sample memory usage
                    current_memory = psutil.Process().memory_info().rss / 1024 / 1024
                    memory_samples.append(current_memory)
                    
                    if iteration % 5 == 0:
                        print(f"Iteration {iteration}: {current_memory:.2f} MB")
                        
                # Analyze memory trend
                if len(memory_samples) > 10:
                    # Check if memory is consistently increasing (potential leak)
                    recent_avg = statistics.mean(memory_samples[-5:])
                    early_avg = statistics.mean(memory_samples[:5])
                    memory_increase = recent_avg - early_avg
                    
                    print(f"Memory increase over iterations: {memory_increase:.2f} MB")
                    
                    # Warn if significant memory increase detected
                    if memory_increase > 50:  # More than 50MB increase
                        print("WARNING: Potential memory leak detected!")
                        
                success = True
                error_msg = ""
                
            except Exception as e:
                success = False
                error_msg = str(e)
                
        profile = profiler.get_profile(success, error_msg)
        self.memory_profiles.append(profile)
        
    def test_concurrent_large_file_processing(self):
        """Test concurrent processing of large files"""
        import concurrent.futures
        
        with MemoryProfiler("concurrent_large_file_processing") as profiler:
            try:
                # Create multiple large files
                file_paths = []
                for i in range(3):
                    file_path = self.create_large_excel(10000, 15)
                    file_paths.append(file_path)
                    
                total_size = sum(self.get_file_size_mb(f) for f in file_paths)
                print(f"Total file size: {total_size:.2f} MB")
                
                # Process files concurrently
                with concurrent.futures.ThreadPoolExecutor(max_workers=3) as executor:
                    futures = []
                    for file_path in file_paths:
                        future = executor.submit(self.extractors['excel'].extract_from_file, file_path)
                        futures.append(future)
                        
                    results = []
                    for future in concurrent.futures.as_completed(futures):
                        result = future.result()
                        results.append(result)
                        
                success = len(results) == len(file_paths)
                error_msg = ""
                
            except Exception as e:
                success = False
                error_msg = str(e)
                
        profile = profiler.get_profile(success, error_msg)
        self.memory_profiles.append(profile)
        
        if success:
            print(f"Concurrent processing: {profile.peak_delta_mb:.2f} MB peak memory")
        else:
            print(f"Concurrent test failed: {error_msg}")
            
    def test_system_resource_limits(self):
        """Test behavior at system resource limits"""
        with MemoryProfiler("system_resource_limits") as profiler:
            try:
                # Get system memory info
                memory_info = psutil.virtual_memory()
                available_mb = memory_info.available / 1024 / 1024
                
                print(f"Available memory: {available_mb:.2f} MB")
                
                # Try to process a file that would use significant memory
                # But don't actually exhaust system memory
                max_safe_size = min(available_mb * 0.1, 100)  # 10% of available or 100MB max
                
                if max_safe_size > 50:
                    # Calculate approximate rows/cols for target size
                    target_rows = int(max_safe_size * 1000)  # Rough estimate
                    file_path = self.create_large_excel(target_rows, 10)
                    
                    actual_size = self.get_file_size_mb(file_path)
                    print(f"Testing with {actual_size:.2f} MB file")
                    
                    result = self.extractors['excel'].extract_from_file(file_path)
                    success = result is not None
                else:
                    print("Skipping resource limit test - insufficient memory")
                    success = True
                    
                error_msg = ""
                
            except MemoryError:
                success = False
                error_msg = "Memory limit reached"
                print("Memory limit reached during resource test")
            except Exception as e:
                success = False
                error_msg = str(e)
                
        profile = profiler.get_profile(success, error_msg)
        self.memory_profiles.append(profile)
        
    @classmethod
    def generate_memory_report(cls):
        """Generate comprehensive memory usage report"""
        if not cls.memory_profiles:
            return
            
        report_data = {
            'test_suite': 'Large File Handling and Memory Stress Tests',
            'timestamp': time.strftime('%Y-%m-%d %H:%M:%S'),
            'system_info': {
                'total_memory_gb': psutil.virtual_memory().total / 1024 / 1024 / 1024,
                'available_memory_gb': psutil.virtual_memory().available / 1024 / 1024 / 1024,
                'cpu_count': psutil.cpu_count(),
                'platform': sys.platform
            },
            'profiles': [profile.__dict__ for profile in cls.memory_profiles],
            'summary': cls._calculate_memory_summary()
        }
        
        # Save JSON report
        report_path = f"memory_stress_report_{int(time.time())}.json"
        with open(report_path, 'w') as f:
            json.dump(report_data, f, indent=2)
            
        # Generate readable report
        readable_path = f"memory_stress_report_{int(time.time())}.txt"
        with open(readable_path, 'w') as f:
            cls._write_memory_report(f, report_data)
            
        print(f"\nMemory stress test reports generated:")
        print(f"  JSON: {os.path.abspath(report_path)}")
        print(f"  Text: {os.path.abspath(readable_path)}")
        
    @classmethod
    def _calculate_memory_summary(cls) -> Dict[str, Any]:
        """Calculate memory usage summary statistics"""
        successful_profiles = [p for p in cls.memory_profiles if p.success]
        
        if not successful_profiles:
            return {}
            
        peak_deltas = [p.peak_delta_mb for p in successful_profiles]
        memory_deltas = [p.memory_delta_mb for p in successful_profiles]
        execution_times = [p.execution_time_sec for p in successful_profiles]
        
        return {
            'total_tests': len(cls.memory_profiles),
            'successful_tests': len(successful_profiles),
            'failed_tests': len(cls.memory_profiles) - len(successful_profiles),
            'avg_peak_memory_mb': statistics.mean(peak_deltas),
            'max_peak_memory_mb': max(peak_deltas),
            'avg_memory_delta_mb': statistics.mean(memory_deltas),
            'max_memory_delta_mb': max(memory_deltas),
            'avg_execution_time_sec': statistics.mean(execution_times),
            'max_execution_time_sec': max(execution_times),
            'total_gc_collections': sum(p.gc_collections for p in successful_profiles)
        }
        
    @classmethod
    def _write_memory_report(cls, f, report_data):
        """Write human-readable memory report"""
        f.write("LARGE FILE HANDLING AND MEMORY STRESS TEST REPORT\n")
        f.write("=" * 60 + "\n\n")
        
        f.write(f"Test Suite: {report_data['test_suite']}\n")
        f.write(f"Timestamp: {report_data['timestamp']}\n\n")
        
        # System information
        f.write("SYSTEM INFORMATION\n")
        f.write("-" * 30 + "\n")
        sys_info = report_data['system_info']
        f.write(f"Total Memory: {sys_info['total_memory_gb']:.2f} GB\n")
        f.write(f"Available Memory: {sys_info['available_memory_gb']:.2f} GB\n")
        f.write(f"CPU Count: {sys_info['cpu_count']}\n")
        f.write(f"Platform: {sys_info['platform']}\n\n")
        
        # Summary statistics
        if report_data['summary']:
            f.write("SUMMARY STATISTICS\n")
            f.write("-" * 30 + "\n")
            summary = report_data['summary']
            f.write(f"Total Tests: {summary['total_tests']}\n")
            f.write(f"Successful: {summary['successful_tests']}\n")
            f.write(f"Failed: {summary['failed_tests']}\n")
            f.write(f"Average Peak Memory: {summary['avg_peak_memory_mb']:.2f} MB\n")
            f.write(f"Maximum Peak Memory: {summary['max_peak_memory_mb']:.2f} MB\n")
            f.write(f"Average Memory Delta: {summary['avg_memory_delta_mb']:.2f} MB\n")
            f.write(f"Maximum Memory Delta: {summary['max_memory_delta_mb']:.2f} MB\n")
            f.write(f"Average Execution Time: {summary['avg_execution_time_sec']:.2f} seconds\n")
            f.write(f"Maximum Execution Time: {summary['max_execution_time_sec']:.2f} seconds\n")
            f.write(f"Total GC Collections: {summary['total_gc_collections']}\n\n")
        
        # Individual test results
        f.write("INDIVIDUAL TEST RESULTS\n")
        f.write("-" * 30 + "\n")
        
        for profile in report_data['profiles']:
            f.write(f"\nTest: {profile['test_name']}\n")
            f.write(f"  Status: {'PASS' if profile['success'] else 'FAIL'}\n")
            if not profile['success']:
                f.write(f"  Error: {profile['error_message']}\n")
            f.write(f"  Execution Time: {profile['execution_time_sec']:.2f} seconds\n")
            f.write(f"  Initial Memory: {profile['initial_memory_mb']:.2f} MB\n")
            f.write(f"  Peak Memory: {profile['peak_memory_mb']:.2f} MB\n")
            f.write(f"  Final Memory: {profile['final_memory_mb']:.2f} MB\n")
            f.write(f"  Memory Delta: {profile['memory_delta_mb']:.2f} MB\n")
            f.write(f"  Peak Delta: {profile['peak_delta_mb']:.2f} MB\n")
            f.write(f"  GC Collections: {profile['gc_collections']}\n")

if __name__ == '__main__':
    # Create test directory
    os.makedirs('tests/stress', exist_ok=True)
    
    # Run stress tests with detailed output
    unittest.main(verbosity=2)